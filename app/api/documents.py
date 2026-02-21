"""
/api/v1/documents endpoints.
Handles upload, listing, detail, and transaction retrieval.
"""

import uuid
from typing import Optional

import structlog
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.dependencies import get_db, get_artifact_store, verify_api_key
from app.models.tables import Document, DocumentSegment, ExtractionRun, Transaction
from app.schemas.documents import (
    DocumentDetail,
    DocumentListResponse,
    DocumentSummary,
    DocumentUploadResponse,
)
from app.schemas.transactions import TransactionListResponse, TransactionResponse
from app.storage.artifact_store import ArtifactStore
from app.storage.paths import doc_hash, raw_pdf_path

logger = structlog.get_logger(__name__)

router = APIRouter(prefix="/api/v1/documents", tags=["documents"], dependencies=[Depends(verify_api_key)])


@router.post("", response_model=DocumentUploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_document(
    file: UploadFile = File(...),
    callback_url: Optional[str] = Query(None),
    priority: int = Query(5, ge=1, le=10),
    metadata_json: Optional[str] = Query(None),
    session: AsyncSession = Depends(get_db),
    store: ArtifactStore = Depends(get_artifact_store),
):
    """Upload a PDF document for extraction."""
    # Validate file type
    if file.content_type not in settings.ALLOWED_MIME_TYPES.split(","):
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"Unsupported file type: {file.content_type}. Allowed: {settings.ALLOWED_MIME_TYPES}",
        )

    # Read file
    file_bytes = await file.read()
    file_size = len(file_bytes)

    # Validate size
    max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if file_size > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large: {file_size} bytes. Max: {max_bytes} bytes",
        )

    # Validate not empty
    if file_size == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Empty file uploaded",
        )

    # Compute hash
    file_hash = doc_hash(file_bytes)
    new_doc_id = str(uuid.uuid4())

    # Store raw file
    raw_path = raw_pdf_path(new_doc_id, file.filename or "document.pdf")
    store.save_bytes(raw_path, file_bytes)

    # Parse metadata if provided
    meta = None
    if metadata_json:
        import json
        try:
            meta = json.loads(metadata_json)
        except json.JSONDecodeError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid JSON in metadata_json parameter",
            )

    # Create DB record
    doc = Document(
        doc_id=uuid.UUID(new_doc_id),
        doc_hash=file_hash,
        file_name=file.filename or "document.pdf",
        file_size_bytes=file_size,
        mime_type=file.content_type or "application/pdf",
        raw_file_uri=raw_path,
        status="QUEUED",
        callback_url=callback_url,
        priority=priority,
        metadata_json=meta,
    )
    session.add(doc)
    await session.flush()

    # Enqueue for background processing
    try:
        from app.worker.jobs import enqueue_extraction
        enqueue_extraction(new_doc_id, priority)
    except Exception as enqueue_err:
        # Redis unavailable — log but don't fail the upload
        logger.warning("enqueue_failed", doc_id=new_doc_id, error=str(enqueue_err))

    logger.info(
        "document_uploaded",
        doc_id=new_doc_id,
        file_name=file.filename,
        file_size_bytes=file_size,
        doc_hash=file_hash,
    )

    return DocumentUploadResponse(
        doc_id=new_doc_id,
        file_name=file.filename or "document.pdf",
        file_size_bytes=file_size,
        doc_hash=file_hash,
        status="QUEUED",
    )


@router.get("", response_model=DocumentListResponse)
async def list_documents(
    status_filter: Optional[str] = Query(None, alias="status"),
    doc_family: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    session: AsyncSession = Depends(get_db),
):
    """List documents with optional filtering and pagination."""
    query = select(Document)

    if status_filter:
        query = query.where(Document.status == status_filter)
    if doc_family:
        query = query.where(Document.doc_family == doc_family)

    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total = (await session.execute(count_query)).scalar() or 0

    # Fetch page
    query = query.order_by(Document.created_at.desc()).offset(offset).limit(limit)
    result = await session.execute(query)
    docs = result.scalars().all()

    return DocumentListResponse(
        documents=[
            DocumentSummary(
                doc_id=str(d.doc_id),
                file_name=d.file_name,
                file_size_bytes=d.file_size_bytes,
                doc_family=d.doc_family,
                provider_guess=d.provider_guess,
                status=d.status,
                page_count=d.page_count,
                created_at=d.created_at,
                updated_at=d.updated_at,
            )
            for d in docs
        ],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.post("/process-all", status_code=status.HTTP_202_ACCEPTED)
async def process_all_queued(
    session: AsyncSession = Depends(get_db),
):
    """
    Process all QUEUED or FAILED documents inline.
    Marks each as PROCESSING first to prevent duplicate runs.
    """
    from app.pipeline.orchestrator import DocumentPipeline

    result = await session.execute(
        select(Document)
        .where(Document.status.in_(["QUEUED", "FAILED"]))
        .order_by(Document.created_at)
    )
    docs = result.scalars().all()

    if not docs:
        return {"message": "No documents to process", "processed": 0}

    # Mark all as PROCESSING to prevent concurrent runs
    seen_ids = set()
    unique_docs = []
    for doc in docs:
        doc_id_str = str(doc.doc_id)
        if doc_id_str not in seen_ids:
            seen_ids.add(doc_id_str)
            doc.status = "RENDERING"
            unique_docs.append(doc)
    await session.commit()

    pipeline = DocumentPipeline()
    results = []

    for doc in unique_docs:
        try:
            output = await pipeline.process(str(doc.doc_id))
            results.append({"doc_id": str(doc.doc_id), "status": output.get("status", "UNKNOWN")})
        except Exception as e:
            results.append({"doc_id": str(doc.doc_id), "status": "FAILED", "error": str(e)})

    return {"processed": len(results), "results": results}


@router.delete("/queue/clear", status_code=status.HTTP_200_OK)
async def clear_queue(
    session: AsyncSession = Depends(get_db),
):
    """Delete all QUEUED documents and their related data (cascading)."""
    result = await session.execute(
        select(Document).where(Document.status == "QUEUED")
    )
    docs = result.scalars().all()

    if not docs:
        return {"message": "No queued documents to clear", "deleted": 0}

    count = len(docs)
    for doc in docs:
        await session.delete(doc)
    await session.flush()

    logger.info("queue_cleared", deleted=count)
    return {"message": f"Cleared {count} queued documents", "deleted": count}


@router.get("/export/all-csv")
async def export_all_transactions_csv(
    session: AsyncSession = Depends(get_db),
):
    """Export all transactions across all documents as a single CSV."""
    import csv
    import io
    from fastapi.responses import StreamingResponse

    result = await session.execute(
        select(Transaction, Document.file_name)
        .join(Document, Transaction.doc_id == Document.doc_id)
        .order_by(Document.file_name, Transaction.row_index)
    )
    rows = result.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "File", "Date", "Description", "Amount", "Direction", "Balance",
        "Direction Source", "Confidence", "Balance Confirmed"
    ])

    for t, fname in rows:
        writer.writerow([
            fname or "",
            str(t.posted_date) if t.posted_date else "",
            t.description_raw or t.description_clean or "",
            str(t.amount) if t.amount else "",
            t.direction or "",
            str(t.running_balance) if t.running_balance else "",
            t.direction_source or "",
            str(t.confidence_overall) if t.confidence_overall else "",
            "Yes" if t.balance_confirmed else "No",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="all_transactions.csv"'},
    )


@router.get("/export/all-xlsx")
async def export_all_transactions_xlsx(
    session: AsyncSession = Depends(get_db),
):
    """Export all transactions across all documents as a formatted Excel file."""
    return await _build_xlsx_response(session, doc_ids=None, filename="all_transactions.xlsx")


@router.get("/export/batch-xlsx")
async def export_batch_transactions_xlsx(
    doc_ids: str = Query(..., description="Comma-separated document IDs"),
    session: AsyncSession = Depends(get_db),
):
    """Export transactions for specific documents as a formatted Excel file."""
    id_list = [d.strip() for d in doc_ids.split(",") if d.strip()]
    if not id_list:
        raise HTTPException(status_code=400, detail="No document IDs provided")

    try:
        uuid_list = [uuid.UUID(d) for d in id_list]
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid document ID format")

    return await _build_xlsx_response(session, doc_ids=uuid_list, filename="batch_transactions.xlsx")


@router.get("/{doc_id}", response_model=DocumentDetail)
async def get_document(
    doc_id: str,
    session: AsyncSession = Depends(get_db),
):
    """Get full document detail including segments and extraction runs."""
    try:
        doc_uuid = uuid.UUID(doc_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid doc_id format")

    result = await session.execute(
        select(Document).where(Document.doc_id == doc_uuid)
    )
    doc = result.scalar_one_or_none()

    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    # Load segments
    seg_result = await session.execute(
        select(DocumentSegment)
        .where(DocumentSegment.doc_id == doc_uuid)
        .order_by(DocumentSegment.segment_index)
    )
    segments = seg_result.scalars().all()

    # Load extraction runs
    run_result = await session.execute(
        select(ExtractionRun)
        .where(ExtractionRun.doc_id == doc_uuid)
        .order_by(ExtractionRun.created_at.desc())
    )
    runs = run_result.scalars().all()

    return DocumentDetail(
        doc_id=str(doc.doc_id),
        doc_hash=doc.doc_hash,
        file_name=doc.file_name,
        file_size_bytes=doc.file_size_bytes,
        mime_type=doc.mime_type,
        page_count=doc.page_count,
        doc_family=doc.doc_family,
        doc_family_confidence=float(doc.doc_family_confidence) if doc.doc_family_confidence else None,
        provider_guess=doc.provider_guess,
        provider_confidence=float(doc.provider_confidence) if doc.provider_confidence else None,
        currency=doc.currency,
        status=doc.status,
        priority=doc.priority,
        metadata_json=doc.metadata_json,
        created_at=doc.created_at,
        updated_at=doc.updated_at,
        segments=[],  # TODO: map segment summaries with tx counts
        extraction_runs=[],  # TODO: map run summaries
    )


@router.delete("/{doc_id}", status_code=status.HTTP_200_OK)
async def delete_document(
    doc_id: str,
    session: AsyncSession = Depends(get_db),
):
    """Delete a document and all related data (segments, transactions, etc.)."""
    try:
        doc_uuid = uuid.UUID(doc_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid doc_id format")

    result = await session.execute(
        select(Document).where(Document.doc_id == doc_uuid)
    )
    doc = result.scalar_one_or_none()

    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    file_name = doc.file_name
    await session.delete(doc)
    await session.flush()

    logger.info("document_deleted", doc_id=doc_id, file_name=file_name)
    return {"message": f"Deleted document: {file_name}", "doc_id": doc_id}


@router.get("/{doc_id}/transactions", response_model=TransactionListResponse)
async def get_document_transactions(
    doc_id: str,
    run_id: Optional[str] = Query(None, description="Specific run ID, defaults to latest"),
    session: AsyncSession = Depends(get_db),
):
    """Get extracted transactions for a document."""
    try:
        doc_uuid = uuid.UUID(doc_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid doc_id format")

    # Find the run to use
    if run_id:
        try:
            run_uuid = uuid.UUID(run_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid run_id format")
    else:
        # Get latest run
        run_result = await session.execute(
            select(ExtractionRun)
            .where(ExtractionRun.doc_id == doc_uuid, ExtractionRun.is_latest == True)
            .limit(1)
        )
        latest_run = run_result.scalar_one_or_none()
        if not latest_run:
            raise HTTPException(status_code=404, detail="No extraction run found for this document")
        run_uuid = latest_run.run_id

    # Fetch transactions
    tx_result = await session.execute(
        select(Transaction)
        .where(Transaction.run_id == run_uuid)
        .order_by(Transaction.page_index, Transaction.row_index)
    )
    txns = tx_result.scalars().all()

    return TransactionListResponse(
        transactions=[
            TransactionResponse(
                transaction_id=str(t.transaction_id),
                page_index=t.page_index,
                row_index=t.row_index,
                posted_date=t.posted_date.isoformat() if t.posted_date else None,
                value_date=t.value_date.isoformat() if t.value_date else None,
                description_raw=t.description_raw,
                description_clean=t.description_clean,
                amount=t.amount,
                currency=t.currency,
                direction=t.direction,
                direction_source=t.direction_source,
                running_balance=t.running_balance,
                balance_confirmed=t.balance_confirmed,
                reference=t.reference,
                transaction_type=t.transaction_type,
                is_balance_marker=t.is_balance_marker,
                confidence_overall=float(t.confidence_overall) if t.confidence_overall else None,
                confidence_amount=float(t.confidence_amount) if t.confidence_amount else None,
                confidence_direction=float(t.confidence_direction) if t.confidence_direction else None,
                confidence_date=float(t.confidence_date) if t.confidence_date else None,
            )
            for t in txns
        ],
        total=len(txns),
        run_id=str(run_uuid),
        doc_id=doc_id,
    )


@router.get("/{doc_id}/export/csv")
async def export_transactions_csv(
    doc_id: str,
    session: AsyncSession = Depends(get_db),
):
    """Export document transactions as a downloadable CSV file."""
    import csv
    import io
    from fastapi.responses import StreamingResponse

    try:
        doc_uuid = uuid.UUID(doc_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid document ID")

    # Get document for filename
    result = await session.execute(
        select(Document).where(Document.doc_id == doc_uuid)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    # Get transactions
    result = await session.execute(
        select(Transaction)
        .where(Transaction.doc_id == doc_uuid)
        .order_by(Transaction.row_index)
    )
    txns = result.scalars().all()

    if not txns:
        raise HTTPException(status_code=404, detail="No transactions found for this document")

    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Date", "Description", "Amount", "Direction", "Balance",
        "Direction Source", "Confidence", "Balance Confirmed"
    ])

    for t in txns:
        writer.writerow([
            str(t.posted_date) if t.posted_date else "",
            t.description_raw or t.description_clean or "",
            str(t.amount) if t.amount else "",
            t.direction or "",
            str(t.running_balance) if t.running_balance else "",
            t.direction_source or "",
            str(t.confidence_overall) if t.confidence_overall else "",
            "Yes" if t.balance_confirmed else "No",
        ])

    output.seek(0)
    safe_name = (doc.file_name or "export").replace(".pdf", "").replace(" ", "_")
    filename = f"{safe_name}_transactions.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{doc_id}/export/xlsx")
async def export_transactions_xlsx(
    doc_id: str,
    session: AsyncSession = Depends(get_db),
):
    """Export document transactions as a formatted Excel file."""
    try:
        doc_uuid = uuid.UUID(doc_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid document ID")

    result = await session.execute(
        select(Document).where(Document.doc_id == doc_uuid)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    safe_name = (doc.file_name or "export").replace(".pdf", "").replace(" ", "_")
    return await _build_xlsx_response(session, doc_ids=[doc_uuid], filename=f"{safe_name}_transactions.xlsx")


@router.post("/{doc_id}/process", status_code=status.HTTP_202_ACCEPTED)
async def process_document_now(
    doc_id: str,
    session: AsyncSession = Depends(get_db),
):
    """
    Trigger processing for a single document immediately (inline, no worker needed).
    Use this when the RQ worker isn't running.
    """
    from app.pipeline.orchestrator import DocumentPipeline

    try:
        doc_uuid = uuid.UUID(doc_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid doc_id format")

    # Verify document exists
    result = await session.execute(select(Document).where(Document.doc_id == doc_uuid))
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    # Run pipeline inline
    pipeline = DocumentPipeline()
    try:
        output = await pipeline.process(doc_id)
        return output
    except Exception as e:
        logger.error("inline_processing_failed", doc_id=doc_id, error=str(e))
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")


# ─── Shared XLSX builder ─────────────────────────────────────

async def _build_xlsx_response(
    session: AsyncSession,
    doc_ids: list | None,
    filename: str,
):
    """Build a formatted Excel file from transactions and return as streaming response."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

    # Query transactions
    query = (
        select(Transaction, Document.file_name)
        .join(Document, Transaction.doc_id == Document.doc_id)
    )
    if doc_ids is not None:
        query = query.where(Transaction.doc_id.in_(doc_ids))
    query = query.order_by(Document.file_name, Transaction.row_index)

    result = await session.execute(query)
    rows = result.all()

    if not rows:
        raise HTTPException(status_code=404, detail="No transactions found")

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="D9E2F3"),
    )
    date_font = Font(name="Arial", size=10)
    text_font = Font(name="Arial", size=10)
    money_font = Font(name="Arial", size=10)
    debit_font = Font(name="Arial", size=10, color="CC0000")
    credit_font = Font(name="Arial", size=10, color="006600")

    # Determine if multi-file
    unique_files = set(fname for _, fname in rows)
    multi_file = len(unique_files) > 1

    # Headers
    headers = []
    if multi_file:
        headers.append("File")
    headers.extend(["Date", "Description", "Amount", "Direction", "Balance"])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Freeze top row
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, (t, fname) in enumerate(rows, 2):
        col = 1
        if multi_file:
            c = ws.cell(row=row_idx, column=col, value=fname or "")
            c.font = text_font
            c.border = thin_border
            col += 1

        # Date
        c = ws.cell(row=row_idx, column=col, value=t.posted_date)
        c.font = date_font
        c.number_format = "DD/MM/YYYY"
        c.border = thin_border
        col += 1

        # Description
        desc = t.description_clean or t.description_raw or ""
        c = ws.cell(row=row_idx, column=col, value=desc)
        c.font = text_font
        c.border = thin_border
        col += 1

        # Amount (signed: negative for debits)
        amt = float(t.amount) if t.amount is not None else None
        if amt is not None and t.direction == "DEBIT":
            amt = -abs(amt)
        c = ws.cell(row=row_idx, column=col, value=amt)
        c.number_format = '£#,##0.00;[Red]-£#,##0.00;"-"'
        c.font = debit_font if t.direction == "DEBIT" else credit_font if t.direction == "CREDIT" else money_font
        c.border = thin_border
        col += 1

        # Direction
        c = ws.cell(row=row_idx, column=col, value=t.direction or "")
        c.font = debit_font if t.direction == "DEBIT" else credit_font if t.direction == "CREDIT" else text_font
        c.border = thin_border
        col += 1

        # Balance
        bal = float(t.running_balance) if t.running_balance is not None else None
        c = ws.cell(row=row_idx, column=col, value=bal)
        c.number_format = '£#,##0.00;[Red]-£#,##0.00;"-"'
        c.font = money_font
        c.border = thin_border

    # Auto-size columns
    col_widths = {"File": 30, "Date": 14, "Description": 50, "Amount": 16, "Direction": 12, "Balance": 16}
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(header, 15)

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )